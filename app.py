import os
import re
import time
import threading
from io import BytesIO
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook
from flask import Flask, jsonify, request, make_response
from flask_cors import CORS

from flask_limiter import Limiter
from flask_limiter.util import get_remote_address


app = Flask(__name__)

# ── CORS / Origin allowlist ──────────────────────────────────────────────────
_frontend_origins = [
    o.strip()
    for o in os.getenv("FRONTEND_ORIGINS", "").split(",")
    if o.strip()
]

if not _frontend_origins:
    CORS(app)
else:
    CORS(app, resources={r"/api/*": {"origins": _frontend_origins}})

ENFORCE_ORIGIN = os.getenv("ENFORCE_ORIGIN", "true").lower() == "true"
ALLOWED_ORIGINS = {o.lower() for o in _frontend_origins}

API_REFRESH_TOKEN = os.getenv("API_REFRESH_TOKEN", "")

def _normalize_origin(value: str) -> str:
    try:
        p = urlparse(value)
        if p.scheme and p.netloc:
            return f"{p.scheme.lower()}://{p.netloc.lower()}"
    except Exception:
        pass
    return ""

def _origin_allowed() -> bool:
    if not ENFORCE_ORIGIN or not ALLOWED_ORIGINS:
        return True

    origin = request.headers.get("Origin", "")
    ref = request.headers.get("Referer", "")

    # Browser fetch sends Origin for cross-origin calls
    if origin:
        return _normalize_origin(origin) in ALLOWED_ORIGINS
    if ref:
        return _normalize_origin(ref) in ALLOWED_ORIGINS

    # If no Origin/Referer, block only if you want to be strict.
    # For SmartEdit / browsers, you should normally get Origin.
    return False

@app.before_request
def _block_unknown_origins():
    if request.path == "/api/healthz":
        return

    # allow refresh with bearer token even if origin is not present
    if request.path == "/api/refresh":
        auth = request.headers.get("Authorization", "")
        if API_REFRESH_TOKEN and auth == f"Bearer {API_REFRESH_TOKEN}":
            return

    if request.path.startswith("/api/"):
        if not _origin_allowed():
            return jsonify({"error": "origin not allowed"}), 403

@app.after_request
def _security_headers(resp):
    resp.headers["Cache-Control"] = "no-store"
    resp.headers["X-Content-Type-Options"] = "nosniff"
    resp.headers["X-Frame-Options"] = "DENY"
    resp.headers["Referrer-Policy"] = "no-referrer"
    resp.headers["Content-Security-Policy"] = (
        "default-src 'none'; frame-ancestors 'none'; base-uri 'none'"
    )
    resp.headers["Strict-Transport-Security"] = (
        "max-age=31536000; includeSubDomains"
    )
    return resp

# ── Rate limiting ────────────────────────────────────────────────────────────
limiter = Limiter(
    key_func=get_remote_address,
    default_limits=[os.getenv("RATE_LIMIT_DEFAULT", "200 per hour")]
)
limiter.init_app(app)

RATE_LIMIT_META = os.getenv("RATE_LIMIT_META", "30/minute;1000/day")
RATE_LIMIT_SUGGEST = os.getenv("RATE_LIMIT_SUGGEST", "60/minute;1500/day")
RATE_LIMIT_SEARCH = os.getenv("RATE_LIMIT_SEARCH", "20/minute;500/day")
RATE_LIMIT_REFRESH = os.getenv("RATE_LIMIT_REFRESH", "5/hour;20/day")

# ── GitHub data source config ────────────────────────────────────────────────
GITHUB_TOKEN = os.getenv("GITHUB_TOKEN", "")
DATA_REPO = os.getenv("DATA_REPO", "")          # "owner/repo"
DATA_PATH = os.getenv("DATA_PATH", "")          # "file.xlsx" or "folder/file.xlsx"
DATA_SHEET = os.getenv("DATA_SHEET", "")        # "Export Worksheet"
DATA_REF = os.getenv("DATA_REF", "main")        # branch/tag/commit

DATA_CACHE_TTL_SEC = int(os.getenv("DATA_CACHE_TTL_SEC", "600"))  # 10 min

SUGGEST_LIMIT = int(os.getenv("SUGGEST_LIMIT", "20"))

# ── In-memory cache ──────────────────────────────────────────────────────────
_lock = threading.Lock()
_cache = {
    "loaded_at": 0.0,
    "rows": [],
    "by_sku": {},          # sku -> row
    "by_label": {},        # label_lower -> row
    "by_core": {},         # core_sku -> [rows] (accessories)
    "last_error": "",
}

def _gh_headers_raw():
    if not GITHUB_TOKEN:
        return {}
    # Fine-grained tokens work with Bearer; classic PAT also works
    return {
        "Authorization": f"Bearer {GITHUB_TOKEN}",
        "Accept": "application/vnd.github.raw",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "accessory-tool-backend",
    }

def _download_excel_bytes() -> bytes:
    if not (DATA_REPO and DATA_PATH and GITHUB_TOKEN):
        raise RuntimeError("Missing DATA_REPO / DATA_PATH / GITHUB_TOKEN")

    # This endpoint returns RAW bytes when Accept: application/vnd.github.raw is used
    url = f"https://api.github.com/repos/{DATA_REPO}/contents/{DATA_PATH}"
    params = {"ref": DATA_REF} if DATA_REF else None

    r = requests.get(url, headers=_gh_headers_raw(), params=params, timeout=30)
    if r.status_code == 404:
        raise RuntimeError("GitHub file not found (check DATA_REPO/DATA_PATH/DATA_REF)")
    if r.status_code == 401 or r.status_code == 403:
        raise RuntimeError("GitHub unauthorized/forbidden (check token permissions)")
    r.raise_for_status()
    return r.content

def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())

def _pick_col(headers_norm: list[str], candidates: list[str]) -> int | None:
    """
    Find a column index by matching normalized header names.
    candidates should already be normalized-like (we normalize inside).
    """
    cand = {_norm(x) for x in candidates}
    for i, h in enumerate(headers_norm):
        if h in cand:
            return i
    return None

def _tokenize(q: str) -> list[str]:
    q = (q or "").strip().lower()
    toks = re.findall(r"[a-z0-9]+", q)
    # de-dup keep order
    seen = set()
    out = []
    for t in toks:
        if t and t not in seen:
            seen.add(t)
            out.append(t)
    return out

def _score(text: str, tokens: list[str]) -> tuple[int, int]:
    """
    Sort key: (matched_token_count, quality)
    More tokens matched => higher rank, as you requested.
    """
    t = text.lower()
    matched = 0
    quality = 0
    for tok in tokens:
        pos = t.find(tok)
        if pos == -1:
            continue
        matched += 1
        quality += 10 + min(len(tok), 10)
        if pos == 0:
            quality += 3
        elif pos < 20:
            quality += 1
    return matched, quality

def _build_label(brand: str, name: str, sku: str) -> str:
    brand = str(brand or "").strip()
    name = str(name or "").strip()
    sku = str(sku or "").strip()

    if brand and name:
        # avoid "brand brand ..."
        if _norm(name).startswith(_norm(brand)):
            label = name
        else:
            label = f"{brand} {name}"
    else:
        label = name or brand or sku

    return label.strip()

def _load_excel_into_cache(force: bool = False):
    with _lock:
        now = time.time()
        if not force and _cache["rows"] and (now - _cache["loaded_at"] < DATA_CACHE_TTL_SEC):
            return

        try:
            b = _download_excel_bytes()
            wb = load_workbook(filename=BytesIO(b), read_only=True, data_only=True)

            sheet_name = DATA_SHEET if DATA_SHEET in wb.sheetnames else wb.sheetnames[0]
            ws = wb[sheet_name]

            # read header row
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if not header:
                raise RuntimeError("Excel sheet has no header row")

            headers = [str(h or "").strip() for h in header]
            headers_norm = [_norm(h) for h in headers]

            # Common column name candidates (adjust if your Excel uses different headers)
            idx_sku = _pick_col(headers_norm, ["Item", "SKU", "Item No", "item", "item_no", "product_number"])
            idx_brand = _pick_col(headers_norm, ["Brand", "brand"])
            idx_dept = _pick_col(headers_norm, ["Department", "department", "Category", "category"])
            idx_name = _pick_col(headers_norm, ["Item Name (retek)", "item_name_retek", "Item Name", "item_name", "Product Name"])
            idx_type = _pick_col(headers_norm, ["item_type", "Item Type", "Type", "type"])
            idx_rrp = _pick_col(headers_norm, ["RRP", "rrp"])
            idx_allow = _pick_col(headers_norm, ["Allow To Buy", "AllowToBuy", "allow_to_buy"])

            # Optional "core link" column (if your sheet has it)
            idx_core = _pick_col(headers_norm, ["Core Item", "core_item", "Parent Item", "parent_item", "Main Item", "main_item", "Core SKU", "core_sku"])

            data_rows = []
            by_sku = {}
            by_label = {}
            by_core = {}

            for r in rows_iter:
                # skip empty rows
                if not r or all(v is None or str(v).strip() == "" for v in r):
                    continue

                def get(i):
                    if i is None:
                        return ""
                    if i >= len(r):
                        return ""
                    v = r[i]
                    return "" if v is None else str(v).strip()

                sku = get(idx_sku)
                brand = get(idx_brand)
                dept = get(idx_dept)
                name = get(idx_name)
                itype = get(idx_type)
                rrp = get(idx_rrp)
                allow = get(idx_allow)
                core = get(idx_core)

                label = _build_label(brand, name, sku)

                row_obj = {
                    "_sku": sku,
                    "_brand": brand,
                    "_dept": dept,
                    "_name": name,
                    "_type": itype,
                    "_rrp": rrp,
                    "_allow": allow,
                    "_core": core,
                    "_label": label,
                    "_search": _norm(f"{sku} {brand} {name} {dept} {itype} {core} {label}"),
                }
                data_rows.append(row_obj)

                if sku:
                    by_sku[sku] = row_obj
                if label:
                    by_label[_norm(label)] = row_obj

                # if accessory rows point to a core SKU, index them
                if core:
                    by_core.setdefault(core, []).append(row_obj)

            _cache["rows"] = data_rows
            _cache["by_sku"] = by_sku
            _cache["by_label"] = by_label
            _cache["by_core"] = by_core
            _cache["loaded_at"] = time.time()
            _cache["last_error"] = ""

        except Exception as e:
            _cache["last_error"] = str(e)
            # Keep old cache if it exists, so service still works with last good data
            if not _cache["rows"]:
                raise

def _ensure_data_loaded():
    _load_excel_into_cache(force=False)

def _result_payload(row_obj):
    # keys that your frontend already knows how to read
    return {
        "department": row_obj.get("_dept", ""),
        "brand": row_obj.get("_brand", ""),
        "item_name_retek": row_obj.get("_label", "") or row_obj.get("_name", ""),
        "item_type": row_obj.get("_type", ""),
        "rrp": row_obj.get("_rrp", ""),
        "allow_to_buy": row_obj.get("_allow", ""),
        "item": row_obj.get("_sku", ""),
    }

def _related_payload(row_obj):
    return {
        "Department": row_obj.get("_dept", ""),
        "Brand": row_obj.get("_brand", ""),
        "Item Name (retek)": row_obj.get("_label", "") or row_obj.get("_name", ""),
        "RRP": row_obj.get("_rrp", ""),
        "Allow To Buy": row_obj.get("_allow", ""),
        "Item": row_obj.get("_sku", ""),
    }

def _is_core(row_obj) -> bool:
    t = (row_obj.get("_type") or "").strip().lower()
    return t.startswith("c") or "core" in t

def _is_accessory(row_obj) -> bool:
    t = (row_obj.get("_type") or "").strip().lower()
    return t.startswith("a") or "access" in t

# ── Routes ───────────────────────────────────────────────────────────────────
@app.get("/api/healthz")
def health():
    return {"ok": True}

@limiter.limit(RATE_LIMIT_META)
@app.get("/api/meta")
def api_meta():
    try:
        _ensure_data_loaded()
    except Exception as e:
        return {
            "ok": False,
            "error": str(e),
            "repo": DATA_REPO,
            "path": DATA_PATH,
            "sheet": DATA_SHEET,
        }, 500

    return {
        "ok": True,
        "loaded_at": _cache["loaded_at"],
        "row_count": len(_cache["rows"]),
        "repo": DATA_REPO,
        "path": DATA_PATH,
        "sheet": DATA_SHEET,
        "ref": DATA_REF,
        "last_error": _cache["last_error"],
    }

@limiter.limit(RATE_LIMIT_SUGGEST)
@app.get("/api/suggest")
def api_suggest():
    q = request.args.get("q", "") or request.args.get("query", "") or request.args.get("term", "")
    q = (q or "").strip()
    if not q:
        return {"suggestions": []}

    try:
        _ensure_data_loaded()
    except Exception as e:
        return jsonify({"error": str(e)}), 502

    tokens = _tokenize(q)
    if not tokens:
        return {"suggestions": []}

    scored = []
    for r in _cache["rows"]:
        matched, quality = _score(r["_search"], tokens)
        if matched >= 1:
            scored.append((matched, quality, r["_label"]))

    # more parts matched => higher rank
    scored.sort(key=lambda x: (x[0], x[1]), reverse=True)

    out = []
    seen = set()
    for _, __, label in scored:
        key = _norm(label)
        if key in seen:
            continue
        seen.add(key)
        out.append(label)
        if len(out) >= SUGGEST_LIMIT:
            break

    return {"suggestions": out}

@limiter.limit(RATE_LIMIT_SEARCH)
@app.post("/api/search")
def api_search():
    body = request.get_json(force=False, silent=True) or {}
    action = (body.get("action") or "").strip().lower()

    try:
        _ensure_data_loaded()
    except Exception as e:
        return jsonify({"error": str(e)}), 502

    def find_by_sku(sku: str):
        return _cache["by_sku"].get(str(sku or "").strip())

    def find_by_label(label: str):
        return _cache["by_label"].get(_norm(label or ""))

    # --- resolve "selected" row ---
    selected = None

    if action == "number":
        selected = find_by_sku(body.get("product_number"))

    elif action == "link":
        link = str(body.get("product_link") or "")
        m = re.search(r"(\d{8})", link)
        if m:
            selected = find_by_sku(m.group(1))

    elif action == "dropdown":
        selected = find_by_label(body.get("selected_item_name"))
        # fallback: broad match if label not found
        if not selected:
            q = str(body.get("selected_item_name") or "")
            tokens = _tokenize(q)
            best = None
            for r in _cache["rows"]:
                matched, quality = _score(r["_search"], tokens)
                if matched < 1:
                    continue
                cand = (matched, quality, r)
                if best is None or (cand[0], cand[1]) > (best[0], best[1]):
                    best = cand
            selected = best[2] if best else None

    else:
        # generic text search
        q = str(body.get("q") or body.get("query") or body.get("term") or "")
        tokens = _tokenize(q)
        best = None
        for r in _cache["rows"]:
            matched, quality = _score(r["_search"], tokens)
            if matched < 1:
                continue
            cand = (matched, quality, r)
            if best is None or (cand[0], cand[1]) > (best[0], best[1]):
                best = cand
        selected = best[2] if best else None

    if not selected:
        return jsonify({"error": "No item found"}), 404

    # --- build related items (only if your Excel has a usable core link column) ---
    related = []
    core_ref = (selected.get("_core") or "").strip()
    sku = (selected.get("_sku") or "").strip()

    if core_ref:
        # if accessory points to a core, return that core
        core_row = find_by_sku(core_ref)
        if core_row:
            related = [_related_payload(core_row)]
    else:
        # if it's a core, return accessories that point to it
        rel_rows = _cache["by_core"].get(sku, [])
        related = [_related_payload(r) for r in rel_rows]

    return {
        "result": _result_payload(selected),
        "related_items": related,
    }

@limiter.limit(RATE_LIMIT_REFRESH)
@app.post("/api/refresh")
def api_refresh():
    if API_REFRESH_TOKEN:
        auth = request.headers.get("Authorization", "")
        if auth != f"Bearer {API_REFRESH_TOKEN}":
            return jsonify({"error": "unauthorized"}), 401

    try:
        _load_excel_into_cache(force=True)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    return {"ok": True, "loaded_at": _cache["loaded_at"], "row_count": len(_cache["rows"])}

@app.errorhandler(429)
def _ratelimit_handler(e):
    return jsonify({"error": "Too many requests. Please try again later."}), 429

if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000"))
    app.run(host="0.0.0.0", port=port, debug=False)
